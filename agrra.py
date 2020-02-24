# agrra.py, by Robert Smiley
# Copyright (c) Global Volunteers International, All Rights Reserved
# Library for loading AGRRA coral transects

from openpyxl import load_workbook
from matplotlib import pyplot
from PIL import Image
from palettable.tableau import Tableau_20 as pallete
import atexit, config, csv, datetime, io, os, random, sqlite3

DATAMAP_PATH = config.APP_PATH + 'datamap/'
SHEET_DATAMAP = DATAMAP_PATH + 'sheet.csv'
TRANSECT_DATAMAP=DATAMAP_PATH + 'transect.csv'
ENCOUNTER_DATAMAP=DATAMAP_PATH + 'encounter.csv'
DB_INIT = config.APP_PATH + 'agrra.sql'
CORAL_INIT = config.APP_PATH + 'coral.csv'

COLORS = pallete.mpl_colors

def gen_config(app_path=config.APP_PATH):
	config_str= \
'''APP_PATH = '{0}'
DB = '{1}'
'''
	with open(config.APP_PATH + 'config.py.tmp', 'w') as config_file:
		config_file.write(config_str.format(app_path, config.DB))
		os.replace(config.APP_PATH + 'config.py.tmp', config.APP_PATH + 'config.py')

def agraa_atexit():
	gen_config()

atexit.register(agraa_atexit)

sheet_datamap = {}
with open(SHEET_DATAMAP) as mapfile:
	reader = csv.DictReader(mapfile)
	for row in reader:
		sheet_datamap[row['key']]=row
		if row['key'] == 'min_row' or row['key'] == 'max_col':
			sheet_datamap[row['key']]['pos'] = int(sheet_datamap[row['key']]['pos']) # convert positional values read to integers

transect_datamap = {}
with open(TRANSECT_DATAMAP) as mapfile:
	reader = csv.DictReader(mapfile)
	for row in reader:
		transect_datamap[row['key']]=row
		transect_datamap[row['key']]['pos']=int(transect_datamap[row['key']]['pos']) # convert positional values read to integers

encounter_datamap = {}
with open(ENCOUNTER_DATAMAP) as mapfile:
	reader = csv.DictReader(mapfile)
	for row in reader:
		encounter_datamap[row['key']]=row
		encounter_datamap[row['key']]['pos']=int(encounter_datamap[row['key']]['pos']) # convert positional values read to integers

db = None
cursor = None

def str2None(x):
	if x == '':
		return None
	else:
		return x

def open_db(name):
	global db
	global cursor

	db = sqlite3.connect(name)

	cursor = db.cursor()

	# Check if table structure has been initialized already. If not, initialize it.
	cursor.execute("SELECT count(name) FROM sqlite_master WHERE type='table' AND name='doc'")
	if not cursor.fetchone()[0] == 1 :
		# Execute db init file...
		with open(DB_INIT) as initfile:
			init = initfile.read()
			for statement in init.split(';'):
				cursor.execute(statement)

		# Initialize coral table...
		coralmap = {}
		with open(CORAL_INIT) as coralfile:
			reader = csv.DictReader(coralfile)
			for i, row in enumerate(reader):
				species_code = None
				if row['species'] != '':
					species_code = row['genus'][0].upper() + row['species'][:3].upper()
				insert_qry = \
'''INSERT INTO coral (coral_id, species_code, genus, species)
VALUES (?, ?, ?, ?);'''
				cursor.execute(insert_qry, [i + 1, species_code, str2None(row['genus']), str2None(row['species'])])

	db.commit()
	config.DB = name

if config.DB:
	open_db(config.DB)


def db_assert():
	global db
	global cursor

	assert db, "No database loaded"
	assert cursor, "No database cursor"

def max_id(table):
	db_assert()
	cursor.execute('SELECT MAX(' + table + '_id) from ' + table)
	mid = cursor.fetchone()[0]
	if not mid:
		mid = 0
	return mid

def sql_insert(table, params, values):
	db_assert()
	values = [str(x) if type(x) == datetime.time else x for x in values]
	params_str = ', '.join(params)
	vals_str = ', '.join(['?' for x in values])
	ins_qry = 'INSERT INTO {0} ({1}) VALUES ({2})'.format(table, params_str, vals_str)
	cursor.execute(ins_qry, values)

def import_xlsx(xlsx):
	db_assert()

	wb = load_workbook(xlsx)
	doc_id = max_id('doc') + 1

	cursor.execute('INSERT INTO doc (doc_id, doc_title) VALUES (?, ?)', [doc_id, os.path.basename(xlsx)])
	for ws in wb:
		sheet_id = str(max_id('sheet') + 1)
		sheet_params = ['sheet_id', 'sheet_title']
		sheet_values = [sheet_id, "'" + ws.title  + "'"]
		for key in sheet_datamap:
			if key != 'min_row' and key != 'max_col':
				sheet_params.append(key)
				sheet_values.append(ws[sheet_datamap[key]['pos']].value)

		sheet_params.append('doc_id')
		sheet_values.append(doc_id)

		sql_insert('sheet', sheet_params, sheet_values)

		trans_id = max_id('transect') + 1
		created_transects = {}
		for row in ws.iter_rows(min_row=sheet_datamap['min_row']['pos'], max_col=sheet_datamap['max_col']['pos']):
			transnum_pos = transect_datamap['transect_num']['pos']
			transnum = row[transnum_pos].value
			if transnum not in created_transects:
				created_transects[transnum] = trans_id
				trans_id += 1
				trans_params = ['transect_id']
				trans_values = [created_transects[transnum]]
				for key in transect_datamap:
					trans_params.append(key)
					trans_values.append(row[transect_datamap[key]['pos']].value)

				trans_params.append('sheet_id')
				trans_values.append(sheet_id)

				sql_insert('transect', trans_params, trans_values)

			enc_id = max_id('encounter') + 1
			enc_params = ['encounter_id']
			enc_values = [enc_id]
			for key in encounter_datamap:
				if key == 'species_code':
					species_code = row[encounter_datamap[key]['pos']].value
					qry = 'SELECT coral_id FROM coral WHERE species_code=?'
					cursor.execute(qry, (species_code,))
					res = cursor.fetchone()
					if not res:
						searchtext = species_code.split()[0] + '%'
						qry = 'SELECT coral_id FROM coral WHERE UPPER(genus) LIKE ? AND species IS NULL'
						cursor.execute(qry, (searchtext,))
						res = cursor.fetchone()
						if not res:
							raise Exception('Unkown Species Code: ' + species_code)
					enc_params.append('coral_id')
					enc_values.append(res[0])
				else:
					enc_params.append(key)
					enc_values.append(row[encounter_datamap[key]['pos']].value)
			enc_params.append('transect_id')
			enc_values.append(created_transects[transnum])
			sql_insert('encounter', enc_params, enc_values)

	db.commit()

def unzip_pairs(lst):
	lst1 = []
	lst2 = []

	for e in lst:
		lst1.append(e[0])
		lst2.append(e[1])

	return lst1, lst2

def piechart(qry, title='', saveas='', interactive=False):
	db_assert()
	cursor.execute(qry)
	res = cursor.fetchall()
	res.sort(key=lambda e: e[1], reverse=True)
	labels, data = unzip_pairs(res)

	total = sum(data)
	percent = [(x / total) * 100 for x in data]

	fig, axis = pyplot.subplots()

	axis.set_prop_cycle(color=COLORS)

	wedges, texts = axis.pie(data, startangle=75, shadow=True)

	chart_labels = []
	for i, label in enumerate(labels):
		chart_labels.append('{0}: {1}% ({2})'.format(label, round(percent[i], 1), data[i]))

	axis.set_title(title)
	axis.axis('equal')

	chartBox = axis.get_position()
	axis.set_position([chartBox.x0, chartBox.y0, chartBox.width*0.6, chartBox.height])
	legend = axis.legend(wedges, chart_labels, title='Total: {0}'.format(total), loc='upper center', bbox_to_anchor=(1.4, 1.15), shadow=True)

	if saveas != '':
		pyplot.savefig(saveas, bbox_extra_artists=(legend,), bbox_inches='tight')
		if interactive:
			img = Image.open(saveas)
			img.show()
	else:
		if interactive:
			buff = io.BytesIO()
			pyplot.savefig(buff, bbox_extra_artists=(legend,), bbox_inches='tight', start_angle=90)
			img = Image.open(buff)
			img.show()
		else:
			raise Exception('No save file specified in non-interactive mode')

