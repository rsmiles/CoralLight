#!/usr/bin/env python3

import csv, os, random, sys
from datetime import datetime
from time import localtime, strftime
from openpyxl import Workbook
from openpyxl import load_workbook

assert len(sys.argv) == 3, 'usage: ./datagen.py site_name output_file'

APP_PATH = os.path.dirname(os.path.abspath(__file__)) + '/'
DATAMAP_PATH = APP_PATH + 'datamap/'
SHEET_DATAMAP = DATAMAP_PATH + 'sheet.csv'
TRANSECT_DATAMAP = DATAMAP_PATH + 'transect.csv'
ENCOUNTER_DATAMAP = DATAMAP_PATH + 'encounter.csv'
CORAL_INIT = APP_PATH + 'coral.csv'
SITE_NAME = sys.argv[1]
DATAFILE = sys.argv[2]
NUM_SHEETS = 2

def isSheetEmpty(ws):
	return ws['A1'].value == None

def writeField(ws, dataMap, key, value):
	ws[dataMap[key]['pos']] = value

def writeRowField(row, dataMap, key, value):
	row[dataMap[key]['pos']].value = value

def createSheet(wb, title):
	global sheet_datamap
	ws = None
	if len(wb.worksheets) == 1:
		if isSheetEmpty(wb.worksheets[0]):
			ws = wb.active
			ws.title = title
		else:
			ws = wb.create_sheet(title=title)
	else:
		ws = wb.create_sheet(title=title)

	sheetTable = {'description': 'AGRRA Coral Data Entry Sheet',
					'version': '5.7: August 2015',
					'copyright': '© Ocean Research & Education Foundation',
					'revision': '2016-05-27',
					'surveyor': 'Test Surveyor',
					'AGRRA_code': 'N/A',
					'site_name': SITE_NAME,
					'date': datetime.today().strftime('%d/%m/%Y'),
					'bottom_temp': 28,
					'bottom_temp_units': 'ºC',
					'instrument_type': 'Dive Computer',
					'level': 'detailed'}

	for key in sheetTable:
		writeField(ws, sheet_datamap, key, sheetTable[key])

	return ws

def createTransectTable(num):
	startDepth = round(random.uniform(3.0, 30.0), 1)
	transectTable = {'transect_num': num,
					'start_time': strftime('%H:%M', localtime()),
					'start_depth': startDepth,
					'end_depth': startDepth - round(random.random(), 1),
					'depth_units': 'm',
					'area_surveyed': 100,
					'area_tallied': 25,
					'transect_comments': ''}
	return transectTable

def createEncounterTable(coralList):
	coral = random.choice(coralList)
	species_code = None

	if coral['species'] == '':
		species_code = coral['genus'].upper() + ' SP.'
	else:
		species_code = coral['genus'][0].upper() + coral['species'][:3].upper()

	length = random.randint(2, 400)

	part_mort_old = random.randint(0, 100)
	part_mort_trans = random.randint(0, 100 - part_mort_old)

	encounterTable = {'species_code': species_code,
						'num_isolates': str(random.randint(1, 11)),
						'length': str(length),
						'width': str(random.randint(1, length - 1)),
						'height': str(random.randint(1, 400)),
						'percent_pale': '',
						'percent_bleached':'',
						'bleach_code': random.choice(('', 'P', 'PB', 'BL')),
						'part_mort_old': str(part_mort_old),
						'part_mort_trans': str(part_mort_trans),
						'part_mort_new': str(random.randint(0, 100 - part_mort_old - part_mort_trans)),
						'disease': random.choice(('', 'SCTLD', 'DS', 'YB', 'WB')),
						'extra_bleach': '',
						'extra_mort': '',
						'comments': random.choice(('', 'SCOR', 'TSOL', 'Orange Icing Sponge', 'PCAR', 'HCAR', 'PFB', 'DFB', 'Outplanted')),
						'point_count_l' :'',
						'point_count_p' :'',
						'point_count_bl': '',
						'point_count_tm': '',
						'point_count_om': '',
						'point_count_other': '',
						'point_count_interval': ''}
	return encounterTable

def createTransect(ws, num, start_row, numEncounters, coralList):
	global MAX_COL
	transectTable = createTransectTable(num)

	for row in ws.iter_rows(min_row = start_row, max_col = MAX_COL, max_row=start_row + numEncounters):
		encounterTable = createEncounterTable(coralList)

		for key in transectTable:
			writeRowField(row, transect_datamap, key, transectTable[key])

		for key in encounterTable:
			writeRowField(row, encounter_datamap, key, encounterTable[key])
	

sheet_datamap = {}
with open(SHEET_DATAMAP) as mapfile:
	reader = csv.DictReader(mapfile)
	for row in reader:
		sheet_datamap[row['key']]=row
		if row['key'] == 'min_row' or row['key'] == 'max_col':
			sheet_datamap[row['key']]['pos'] = int(sheet_datamap[row['key']]['pos']) # convert positional values read to integers

MIN_ROW = sheet_datamap['min_row']['pos']
MAX_COL = sheet_datamap['max_col']['pos']

transect_datamap = {}
with open(TRANSECT_DATAMAP) as mapfile:
	reader = csv.DictReader(mapfile)
	for row in reader:
		transect_datamap[row['key']]=row
		transect_datamap[row['key']]['pos']=int(transect_datamap[row['key']]['pos']) # convert positional values read to integers

encounter_datamap = {}
with open(ENCOUNTER_DATAMAP) as mapfile:
	reader = csv.DictReader(mapfile)
	for row in reader:
		encounter_datamap[row['key']]=row
		encounter_datamap[row['key']]['pos']=int(encounter_datamap[row['key']]['pos']) # convert positional values read to integers

wb = None
if os.path.isfile(DATAFILE):
	wb = load_workbook(DATAFILE)
else:
	wb = Workbook()

sheetNum = 1
transectNum = 1

coralList = []
with open(CORAL_INIT) as coralFile:
	coralList = list(csv.DictReader(coralFile))

numCorals = len(coralList)

while sheetNum <= NUM_SHEETS:
	sheetTitle = 'Transect {0} + {1}'.format(str(transectNum), str(transectNum + 1))

	ws = createSheet(wb, sheetTitle)

	createTransect(ws, transectNum, MIN_ROW, 200, coralList)
	transectNum += 1
	createTransect(ws, transectNum, MIN_ROW + 201, 200, coralList)
	transectNum += 1
	sheetNum += 1

wb.save(DATAFILE)

