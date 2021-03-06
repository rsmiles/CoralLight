"""
backend.py

Copyright (c) 2020 Robert Smiley, all rights reserved.
Contents of this module are available under the terms of the GNU General Public
License, version 3. See LICENSE for details.

Backend code for CoralLight.
It contains functions that implement it's key features:
	-Creation of new databases
	-Connecting to databases
	-Initialization of database tables
	-Importing data from an excel file into the database
	-Generating charts from SQL queries
	-Basic displaying of generated charts
	-Saving generated charts
"""

import matplotlib

# Since we don't use matplotlib for displaying the images it generates, we use
# the Agg backend to avoid conflicting with TkInter.
matplotlib.use('Agg')

from openpyxl import load_workbook
from matplotlib import pyplot
from matplotlib import font_manager
from PIL import Image
from PIL import ImageFont
from PIL import ImageDraw
from palettable.tableau import Tableau_20 as pallete_20
from palettable.tableau import Tableau_10 as pallete_10
import atexit, csv, datetime, io, os, random, sqlite3
from . import config

# Get the path of our application
APP_PATH = os.path.dirname(os.path.abspath(__file__)) + '/'

DATAMAP_PATH = APP_PATH + 'datamap/'
SHEET_DATAMAP = DATAMAP_PATH + 'sheet.csv'
TRANSECT_DATAMAP = DATAMAP_PATH + 'transect.csv'
ENCOUNTER_DATAMAP = DATAMAP_PATH + 'encounter.csv'
DB_INIT = APP_PATH + 'db_init.sql'
CORAL_INIT = APP_PATH + 'coral.csv'

COLORS_10 = pallete_10.mpl_colors
COLORS_20 = pallete_20.mpl_colors

FONT_PATH = font_manager.findfont(font_manager.FontProperties())
FONT_PATH_BOLD = font_manager.findfont(font_manager.FontProperties(weight='bold'))
TITLE_FONT_SIZE = 15

TITLE_FONT_PIL = ImageFont.truetype(FONT_PATH_BOLD, 15)

def gen_config(app_path=APP_PATH):
	config_str= "DB = '{0}'"

	with open(APP_PATH + 'config.py.tmp', 'w') as config_file:
		if config.DB:
			config_file.write(config_str.format(config.DB))
		else:
			config_file.write('DB = None')
		os.replace(APP_PATH + 'config.py.tmp', APP_PATH + 'config.py')

def agraa_atexit():
	gen_config()

atexit.register(agraa_atexit)

# Load "datamaps": csv files specifying where information can be found in the AGRRA excel files.

sheet_datamap = {}
with open(SHEET_DATAMAP) as mapfile:
	reader = csv.DictReader(mapfile)
	for row in reader:
		sheet_datamap[row['key']]=row
		if row['key'] == 'min_row' or row['key'] == 'max_col':
			sheet_datamap[row['key']]['pos'] = int(sheet_datamap[row['key']]['pos']) # convert positional values read to integers

transect_datamap = {}
with open(TRANSECT_DATAMAP) as mapfile:
	reader = csv.DictReader(mapfile)
	for row in reader:
		transect_datamap[row['key']]=row
		transect_datamap[row['key']]['pos']=int(transect_datamap[row['key']]['pos']) # convert positional values read to integers

encounter_datamap = {}
with open(ENCOUNTER_DATAMAP) as mapfile:
	reader = csv.DictReader(mapfile)
	for row in reader:
		encounter_datamap[row['key']]=row
		encounter_datamap[row['key']]['pos']=int(encounter_datamap[row['key']]['pos']) # convert positional values read to integers

db = None
cursor = None

def str2None(x):
	if x == '':
		return None
	else:
		return x

def open_db(name):
	global db
	global cursor

	exists = os.path.isfile(name)

	db = sqlite3.connect(name)

	cursor = db.cursor()

	# Check if table structure has been initialized already. If not, initialize it.
	cursor.execute("SELECT count(name) FROM sqlite_master WHERE type='table' AND name='doc'")
	if not cursor.fetchone()[0] == 1 :
		assert not exists, 'Could not open "{0}": Not a CoralLight database'.format(os.path.basename(name))
		# Execute db init file...
		with open(DB_INIT) as initfile:
			init = initfile.read()
			for statement in init.split(';'):
				cursor.execute(statement)

		# Initialize coral table...
		coralmap = {}
		with open(CORAL_INIT) as coralfile:
			reader = csv.DictReader(coralfile)
			for i, row in enumerate(reader):
				species_code = None
				if row['species'] != '':
					species_code = row['genus'][0].upper() + row['species'][:3].upper()
				insert_qry = \
'''INSERT INTO coral (coral_id, species_code, genus, species)
VALUES (?, ?, ?, ?);'''
				cursor.execute(insert_qry, [i + 1, species_code, str2None(row['genus']), str2None(row['species'])])

	db.commit()
	config.DB = name

if config.DB:
	if os.path.isfile(config.DB):
		open_db(config.DB)
	else:
		config.DB = None


def db_assert():
	global db
	global cursor

	assert db, "No database loaded"
	assert cursor, "No database cursor"

def max_id(table):
	db_assert()
	cursor.execute('SELECT MAX(' + table + '_id) from ' + table)
	mid = cursor.fetchone()[0]
	if not mid:
		mid = 0
	return mid

def sql_insert(table, params, values):
	db_assert()

	values = [str(x) if type(x) == datetime.time else x for x in values]
	params_str = ', '.join(params)
	vals_str = ', '.join(['?' for x in values])
	ins_qry = 'INSERT INTO {0} ({1}) VALUES ({2})'.format(table, params_str, vals_str)
	cursor.execute(ins_qry, values)

def verify_workbook(doc_name, wb):
	for ws in wb:
		assert ws[sheet_datamap['description']['pos']].value == 'AGRRA Coral Data Entry Sheet', \
		'Could not import "{0}": Not an AGRRA Coral Data Entry Sheet'.format(os.path.basename(doc_name))

def import_xlsx(xlsx):
	db_assert()

	wb = load_workbook(xlsx)

	verify_workbook(xlsx, wb)

	doc_id = max_id('doc') + 1

	cursor.execute('INSERT INTO doc (doc_id, doc_title) VALUES (?, ?)', [doc_id, os.path.basename(xlsx)])
	for ws in wb:
		sheet_id = str(max_id('sheet') + 1)
		sheet_params = ['sheet_id', 'sheet_title']
		sheet_values = [sheet_id, "'" + ws.title  + "'"]
		for key in sheet_datamap:
			if key != 'min_row' and key != 'max_col':
				sheet_params.append(key)
				if key == 'date': # re-format date to something SQLite will understand
					val = ws[sheet_datamap[key]['pos']].value
					sheet_values.append('-'.join(reversed(val.split('/'))))
				else:
					sheet_values.append(ws[sheet_datamap[key]['pos']].value)

		sheet_params.append('doc_id')
		sheet_values.append(doc_id)

		sql_insert('sheet', sheet_params, sheet_values)

		trans_id = max_id('transect') + 1
		created_transects = {}
		for row in ws.iter_rows(min_row=sheet_datamap['min_row']['pos'], max_col=sheet_datamap['max_col']['pos']):
			transnum_pos = transect_datamap['transect_num']['pos']
			transnum = row[transnum_pos].value
			if transnum not in created_transects:
				created_transects[transnum] = trans_id
				trans_id += 1
				trans_params = ['transect_id']
				trans_values = [created_transects[transnum]]
				for key in transect_datamap:
					trans_params.append(key)
					trans_values.append(row[transect_datamap[key]['pos']].value)

				trans_params.append('sheet_id')
				trans_values.append(sheet_id)

				sql_insert('transect', trans_params, trans_values)

			enc_id = max_id('encounter') + 1
			enc_params = ['encounter_id']
			enc_values = [enc_id]
			for key in encounter_datamap:
				if key == 'species_code':
					species_code = row[encounter_datamap[key]['pos']].value
					qry = 'SELECT coral_id FROM coral WHERE species_code=?'
					cursor.execute(qry, (species_code,))
					res = cursor.fetchone()
					if not res:
						searchtext = species_code.split()[0] + '%'
						qry = 'SELECT coral_id FROM coral WHERE UPPER(genus) LIKE ? AND species IS NULL'
						cursor.execute(qry, (searchtext,))
						res = cursor.fetchone()
						if not res:
							raise Exception('Unkown Species Code: ' + species_code)
					enc_params.append('coral_id')
					enc_values.append(res[0])
				else:
					enc_params.append(key)
					enc_values.append(row[encounter_datamap[key]['pos']].value)
			enc_params.append('transect_id')
			enc_values.append(created_transects[transnum])
			sql_insert('encounter', enc_params, enc_values)

	db.commit()

def unzip_pairs(lst):
	lst1 = []
	lst2 = []

	for e in lst:
		lst1.append(e[0])
		lst2.append(e[1])

	return lst1, lst2

def chart(qry, chart_type, title='', ymax=None):
	db_assert()
	cursor.execute(qry)
	res = cursor.fetchall()
	res.sort(key=lambda e: e[1], reverse=True)
	labels, data = unzip_pairs(res)
	data = [round(x, 1) for x in data]

	total = round(sum(data), 1)
	npoints = len(data)

	percent = [(x / total) * 100 for x in data]

	fig, axis = pyplot.subplots()

	if npoints > 10:
		axis.set_prop_cycle(color=COLORS_20)
	else:
		axis.set_prop_cycle(color=COLORS_10)

	buff = io.BytesIO()

	img = None
	draw = None

	if chart_type == 'pie':
		wedges, texts = axis.pie(data, startangle=75, shadow=True)
		axis.axis('equal')

		chartBox = axis.get_position()
		axis.set_position([chartBox.x0, chartBox.y0, chartBox.width*0.6, chartBox.height])

		chart_labels = []
		for i, label in enumerate(labels):
			chart_labels.append('{0}: {1}% ({2})'.format(label, round(percent[i], 1), data[i]))

		legend = axis.legend(wedges, chart_labels, title='Total: {0}'.format(total), loc='upper left', bbox_to_anchor=(1.1, 0.9), shadow=True)
		pyplot.savefig(buff, bbox_extra_artists=(legend,), bbox_inches='tight', start_angle=90)
		img = Image.open(buff)
		draw = ImageDraw.Draw(img)

		if title != '':
			imgwidth, imgheight = img.size
			textwidth, textheight = draw.textsize(title, font=TITLE_FONT_PIL)
			text_x = imgwidth / 2 - textwidth / 2
			draw.text((text_x, 10), title, fill='black', font=TITLE_FONT_PIL)

	if chart_type == 'bar':
		if title != '':
			oldWeight = pyplot.rcParams['font.weight']
			oldSize = pyplot.rcParams['font.size']

			pyplot.rcParams['font.weight'] = 'bold'
			pyplot.rcParams['font.size'] = TITLE_FONT_SIZE
			
			pyplot.title(title)

			pyplot.rcParams['font.weight'] = oldWeight
			pyplot.rcParams['font.size'] = oldSize
			

		chart_labels = [str(label) +  ' (' + str(data[i]) + ')' for i, label in enumerate(labels)]
		pyplot.xticks(range(len(chart_labels)), chart_labels, rotation='vertical')

		if ymax:
			axis.set_ylim(ymax=ymax)

		axis.bar(range(len(chart_labels)), data)
		pyplot.savefig(buff, bbox_inches='tight')
		img = Image.open(buff)

	pyplot.close(fig)

	return img

