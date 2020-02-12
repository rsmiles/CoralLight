from openpyxl import load_workbook

def load_sheet(ws):
	data = {}

	data['title'] = ws.title
	data['description'] = ws['A1'].value
	data['version'] = ws['G1'].value
	data['copyright'] = ws['J1'].value
	data['revision'] = ws['N1'].value
	data['surveyor'] = ws['C2'].value
	data['AGRRA_code'] = ws['C3'].value
	data['site_name'] = ws['C4'].value
	data['date'] = ws['C5'].value
	data['bottom_temp'] = ws['I2'].value
	data['instrument_type'] = ws['I4'].value
	data['level'] = ws['I5'].value
	data['site_comments'] = ws['N2'].value

	data['transects'] = []

	for row in ws.iter_rows(min_row=10, max_col=31):
		# if we don't have enough transects in our list to store this number, keep adding them until we have enough
		if row[0].value > len(data['transects']):
			i = len(data['transects'])
			while i < row[0].value:
				data['transects'].append({})
				i += 1
			# add transect info
			trans = data['transects'][row[0].value - 1]
			trans['transect_#'] = row[0].value
			trans['start_time'] = row[1].value
			trans['start_depth'] = row[2].value
			trans['end_depth'] = row[3].value
			trans['depth_units'] = row[4].value
			trans['area_surveyed'] = row[5].value
			trans['area_tallied'] = row[6].value
			trans['encounters'] = []

		enc = {}
		enc['species_code'] = row[7].value
		enc['#_isolates'] = row[8].value
		enc['length'] = row[9].value
		enc['width'] = row[10].value
		enc['height'] = row[11].value
		enc['%_pale'] = row[12].value
		enc['%_bleached'] = row[13].value
		enc['bleach_code'] = row[14].value
		enc['%_pm_new'] = row[15].value
		enc['%_pm_trans'] = row[16].value
		enc['%_pm_old'] = row[17].value
		enc['disease'] = row[18].value
		enc['extra_bleach'] = row[19].value
		enc['extra_mortality'] = row[20].value
		enc['comments'] = row[21].value
		enc['pc_l'] = row[22].value
		enc['pc_p'] = row[23].value
		enc['pc_bl'] = row[24].value
		enc['pc_nm'] = row[25].value
		enc['pc_tm'] = row[26].value
		enc['pc_om'] = row[28].value
		enc['pc_other'] = row[29].value
		enc['transect_comments'] = row[30].value
		data['transects'][row[0].value - 1]['encounters'].append(enc)
		
	return data

def load(xlsx):
	data_sheets = {}
	wb = load_workbook(xlsx)
	for ws in wb:
		data_sheets[ws.title] = load_sheet(ws)
	return data_sheets

